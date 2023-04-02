from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl
from pathlib import Path
from datetime import date
import datetime
from constants import globalConstants
from constants.globalConstants import *

class Test_Multi_Tests:
    
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()
    
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def screenShot(self,username_ss = standart_user,password_ss = secret_sauce):
        self.driver.save_screenshot(f"{self.folderPath}/test-volid-login{username_ss}-{password_ss}.png")
    
    def error_button_click(self):
        error_button_off = self.driver.find_element(By.XPATH,error_button_off_path)
        error_button_off.click()

    def getData():
        excelFile = openpyxl.load_workbook(excel_path)
        selectedSheet = excelFile[sheet_number]

        excel_data = []
        total_Rows = selectedSheet.max_row
        
        for i in range(2,total_Rows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            excel_data.append(tupleData)
        
        return excel_data
    
    def login(self,username,password):
        self.waitForElementVisible((By.ID,username_ID))
        username_Input = self.driver.find_element(By.ID,username_ID)
        self.waitForElementVisible((By.ID,password_ID))
        password_Input = self.driver.find_element(By.ID,password_ID)

        username_Input.send_keys(username)
        password_Input.send_keys(password)

        self.waitForElementVisible((By.ID,loginbutton_ID))
        loginBtn = self.driver.find_element(By.ID,loginbutton_ID)        
        loginBtn.click()
    
    def test_sorting_change(self,username = standart_user,password=secret_sauce):
        self.login(username,password)
        self.waitForElementVisible((By.XPATH,sorting_button_xpath))
        for i in sorting_option:
            sorting_btn = self.driver.find_element(By.XPATH,i)
            sorting_btn.click()
            sleep(2)
            self.driver.save_screenshot(f"{self.folderPath}/test_username_password_{username}-{password}-{datetime.datetime.now().microsecond}.png")
       
    def test_valid_login(self,username = standart_user,password = secret_sauce):
        self.login(username,password)
        self.screenShot(username_ss = standart_user,password_ss = secret_sauce)
  
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.login(username,password)
        self.waitForElementVisible((By.XPATH,error_message_path))
        errorMessage = self.driver.find_element(By.XPATH,error_message_path)
        self.screenShot(username,password)
        assert errorMessage.text == error_text 
        self.error_button_click()
    
    def test_locked_out_user(self,username = locked_out_user,password =secret_sauce ):
        self.login(username,password)
        errorMessage = self.driver.find_element(By.XPATH,error_message_path)
        self.screenShot()
        assert errorMessage.text == error_message_lockedOut
        self.error_button_click()

    def test_problem_user(self,username = problem_user, password = secret_sauce):
        self.login(username,password)
        self.waitForElementVisible((By.XPATH,problem_user_img_xpath))
        self.screenShot(username,password)

    def test_problemuser_shopping(self,username = problem_user,password = secret_sauce):
        self.login(username,password)
        self.waitForElementVisible((By.ID,addToC_backBack_ID))
        addToCart_Button = self.driver.find_element(By.ID,addToC_backBack_ID)
        addToCart_Button.click()
        
        self.waitForElementVisible((By.XPATH,shopping_cart_link_xpath))
        shoppingCart_Button = self.driver.find_element(By.XPATH,shopping_cart_link_xpath)
        shoppingCart_Button.click()

        self.waitForElementVisible((By.ID,checkoutButton_ID))
        checkout_Button = self.driver.find_element(By.ID,checkoutButton_ID)
        checkout_Button.click()

        self.waitForElementVisible((By.XPATH,checkout_info_xpath))
        for i in checkout_page_list:
            checkOut_info= self.driver.find_element(By.ID,i)
            checkOut_info.send_keys("test")
        
        self.waitForElementVisible((By.ID,continueButton_ID))
        continue_Button = self.driver.find_element(By.ID,continueButton_ID)
        continue_Button.click()

        self.waitForElementVisible((By.XPATH,checkout_error_xpath))
        checkoutError_Message = self.driver.find_element(By.XPATH,checkout_error_xpath)
        assert checkoutError_Message.text == checkout_error_message
 
    def test_logout(self,username = standart_user, password =secret_sauce):
        self.login(username,password)
        self.waitForElementVisible((By.ID,menu_button_ID))
        menu_btn = self.driver.find_element(By.ID,menu_button_ID)
        menu_btn.click()
        self.waitForElementVisible((By.ID,logout_bar_ID))
        logout_btn = self.driver.find_element(By.ID,logout_bar_ID)
        self.screenShot()
        logout_btn.click()
        
    def test_goTo_socialMedia(self,username = standart_user ,password = secret_sauce):
        self.login(username,password)
        for i in social_media_links:
            self.waitForElementVisible((By.XPATH,i))
            socialMedia_btn = self.driver.find_element(By.XPATH,i)
            socialMedia_btn.click()
            sleep(2)
    def test_find_list_items(self,username = standart_user,password = secret_sauce):
        self.login(username,password)
        self.waitForElementVisible((By.CLASS_NAME,list_items_ClassName))
        list_items = self.driver.find_elements(By.CLASS_NAME,list_items_ClassName)
        assert len(list_items) == 6
    def test_items_links_click(self,username = standart_user,password = secret_sauce):
        self.login(username,password)
        for i in items_links_xpaths:
            self.waitForElementVisible((By.XPATH,i))
            items_links = self.driver.find_element(By.XPATH,i)
            items_links.click()
            sleep(2)
            self.driver.find_element(By.ID,back_to_products_ID)
            backToProducts_Btn = self.driver.find_element(By.ID,back_to_products_ID)
            backToProducts_Btn.click()
    def test_buy_all_products(self,username = standart_user,password = secret_sauce):
        self.login(username,password)
        self.waitForElementVisible((By.CLASS_NAME,list_ClassName))
        
        for i in list_of_items_ID:
            addToCart = self.driver.find_element(By.ID,i)
            addToCart.click()
            
        
        self.waitForElementVisible((By.XPATH,shopping_cart_link_xpath))
        shopping_cart_link = self.driver.find_element(By.XPATH,shopping_cart_link_xpath)
        shopping_cart_link.click()
        

        self.waitForElementVisible((By.ID,removeButton_bike_light_ID))
        removeButton = self.driver.find_element(By.ID,"remove-sauce-labs-bike-light")
        removeButton.click()
        
        self.waitForElementVisible((By.ID,checkoutButton_ID))
        checkOut_Button = self.driver.find_element(By.ID,checkoutButton_ID)
        checkOut_Button.click()
        
        self.waitForElementVisible((By.XPATH,checkout_info_xpath))
        for j in checkout_page_list:
            chechkOut_info = self.driver.find_element(By.ID,j)
            chechkOut_info.send_keys("test")
            
        
        self.waitForElementVisible((By.ID,continueButton_ID))
        continue_Button = self.driver.find_element(By.ID,continueButton_ID)
        continue_Button.click()  
        sleep(2)

        self.waitForElementVisible((By.ID,finishButton_ID))
        finish_Button = self.driver.find_element(By.ID,finishButton_ID) 
        finish_Button.click()
        
        self.waitForElementVisible((By.ID,backHomeButton_ID))
        backHome_Button = self.driver.find_element(By.ID,backHomeButton_ID)
        backHome_Button.click()   
           

        
        







        
        
        
